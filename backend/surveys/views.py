import csv
import io
import openpyxl
from django.http import HttpResponse
from rest_framework import status, permissions, serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, MultiPartParser
from drf_spectacular.utils import extend_schema, inline_serializer
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db import models
import json
from .permissions import IsSurveyParticipantOrAdmin

from .serializers import (
    SurveyCreateSerializer, SurveyDetailSerializer, SurveyUpdateSerializer,
    QuestionSerializer, QuestionUpdateSerializer,
    SurveyQuestionLinkSerializer, RespondentAnswerCreateSerializer,
    SurveyArchiveSerializer, RespondentSurveyStatusSerializer,
    RespondentAnswerDetailSerializer, SurveyRequiredCharacteristicSerializer
)
from .models import Surveys, Questions, SurveyQuestions, RespondentAnswers, SurveyArchive, RespondentSurveyStatus
from core.models import SurveyRequiredCharacteristics

tag = ['Опросы']


def role_allowed(user, roles):
    return getattr(user, 'role', None) in roles


# ------------------- ОПРОСЫ -------------------

class SurveyCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Создать опрос",
        description="Создание нового опроса (customer/moderator).",
        request=SurveyCreateSerializer,
        responses={201: SurveyDetailSerializer},
        tags=tag
    )
    def post(self, request):
        if not role_allowed(request.user, ['customer', 'moderator']):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        serializer = SurveyCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        survey = serializer.save(creator=request.user, status='draft')
        return Response(SurveyDetailSerializer(survey).data, status=status.HTTP_201_CREATED)


class MySurveysView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Мои опросы", responses={200: SurveyDetailSerializer(many=True)}, tags=tag)
    def get(self, request):
        qs = Surveys.objects.filter(creator=request.user)
        return Response(SurveyDetailSerializer(qs, many=True).data, status=status.HTTP_200_OK)


class SurveyRetrieveUpdateDeleteView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Редактирование опроса", request=SurveyUpdateSerializer,
                   responses={200: SurveyDetailSerializer}, tags=tag)
    def put(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        serializer = SurveyUpdateSerializer(survey, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(SurveyDetailSerializer(survey).data, status=status.HTTP_200_OK)

    @extend_schema(summary="Удаление опроса", responses={204: None}, tags=tag)
    def delete(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        survey.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SurveyArchiveView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Архивация опроса", responses={200: SurveyDetailSerializer}, tags=tag)
    def post(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)

        archive, _ = SurveyArchive.objects.get_or_create(survey=survey)
        survey.status = 'stopped'
        survey.save()
        return Response(SurveyDetailSerializer(survey).data, status=status.HTTP_200_OK)


class ArchivedSurveysListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Список архивированных опросов", responses={200: SurveyArchiveSerializer(many=True)}, tags=tag)
    def get(self, request):
        archives = SurveyArchive.objects.all()
        return Response(SurveyArchiveSerializer(archives, many=True).data, status=status.HTTP_200_OK)


class SurveyRestoreView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Восстановить опрос из архива", responses={200: SurveyDetailSerializer}, tags=tag)
    def post(self, request, archive_id: int):
        archive = get_object_or_404(SurveyArchive, pk=archive_id)
        survey = archive.survey
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)

        archive.delete()
        survey.status = 'draft'
        survey.save()
        return Response(SurveyDetailSerializer(survey).data, status=status.HTTP_200_OK)


# ------------------- ВОПРОСЫ -------------------

class QuestionCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Создать вопрос",
        request=QuestionSerializer,
        responses={201: inline_serializer(
            name='ИдентификаторВопроса',
            fields={'question_id': serializers.IntegerField()}
        )}, tags=tag
    )
    def post(self, request):
        if not role_allowed(request.user, ['customer', 'moderator']):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        serializer = QuestionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        q = serializer.save()
        return Response({'question_id': q.question_id}, status=status.HTTP_201_CREATED)


class QuestionUpdateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Редактирование вопроса", request=QuestionUpdateSerializer,
                   responses={200: QuestionSerializer}, tags=tag)
    def put(self, request, question_id: int):
        question = get_object_or_404(Questions, pk=question_id)
        if not role_allowed(request.user, ['moderator', 'customer']):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        serializer = QuestionUpdateSerializer(question, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(QuestionSerializer(question).data, status=status.HTTP_200_OK)


class SurveyQuestionDeleteView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Удалить вопрос из опроса (по ID самого вопроса)",
        responses={204: None},
        tags=tag
    )
    def delete(self, request, question_id: int):
        question = get_object_or_404(Questions, pk=question_id)
        link = SurveyQuestions.objects.filter(question=question).select_related('survey').first()
        if not link:
            return Response({"detail": "Вопрос не привязан ни к одному опросу"},
                            status=status.HTTP_400_BAD_REQUEST)

        if not (link.survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)

        question.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SurveyQuestionLinkView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Привязать вопрос к опросу",
        request=SurveyQuestionLinkSerializer,
        responses={201: inline_serializer(
            name='ИдентификаторПривязки',
            fields={'survey_question_id': serializers.IntegerField()}
        )}, tags=tag
    )
    def post(self, request):
        if not role_allowed(request.user, ['customer', 'moderator']):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        serializer = SurveyQuestionLinkSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        survey = get_object_or_404(Surveys, pk=serializer.validated_data['survey'].survey_id)
        if survey.creator != request.user and request.user.role != 'moderator':
            return Response({"detail": "Нельзя добавлять вопросы в чужой опрос"},
                            status=status.HTTP_403_FORBIDDEN)
        sq = serializer.save()
        return Response({'survey_question_id': sq.survey_question_id}, status=status.HTTP_201_CREATED)


class SurveyQuestionsListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Список вопросов опроса",
                   responses={200: QuestionSerializer(many=True)}, tags=tag)
    def get(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        qlinks = SurveyQuestions.objects.filter(survey=survey).select_related('question').order_by('order')
        questions = [ql.question for ql in qlinks]
        return Response(QuestionSerializer(questions, many=True).data, status=status.HTTP_200_OK)


# ------------------- ОТВЕТЫ -------------------

class RespondentAnswerView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Отправить ответ",
        request=RespondentAnswerCreateSerializer,
        responses={201: inline_serializer(
            name='ИдентификаторОтвета',
            fields={'answer_id': serializers.IntegerField()}
        )}, tags=tag
    )
    def post(self, request):
        if not role_allowed(request.user, ['respondent']):
            return Response({"detail": "Только респонденты могут отправлять ответы"},
                            status=status.HTTP_403_FORBIDDEN)
        serializer = RespondentAnswerCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        ans = serializer.save()
        return Response({'answer_id': ans.answer_id}, status=status.HTTP_201_CREATED)


class SurveyAnswersView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Все ответы на опрос",
        responses={200: inline_serializer(
            name='ОтветыОпроса',
            fields={'answers': serializers.ListField(child=serializers.DictField())}
        )}, tags=tag
    )
    def get(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        answers = RespondentAnswers.objects.filter(
            survey_question__survey=survey
        ).select_related('respondent', 'survey_question', 'survey_question__question')
        out = []
        for a in answers:
            out.append({
                'answer_id': a.answer_id,
                'survey_question_id': a.survey_question.survey_question_id,
                'question_id': a.survey_question.question.question_id,
                'respondent_id': a.respondent.user_id,
                'text_answer': a.text_answer,
                'created_at': a.created_at,
            })
        return Response({'answers': out}, status=status.HTTP_200_OK)


# ------------------- СТАТУСЫ и ДОСТУПНЫЕ -------------------

class SurveyToggleStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Изменить статус опроса",
        request=inline_serializer(
            name='НовыйСтатус',
            fields={'status': serializers.ChoiceField(
                choices=['draft', 'active', 'stopped', 'finished']
            )}
        ),
        responses={200: inline_serializer(
            name='СтатусОпроса',
            fields={'survey_id': serializers.IntegerField(), 'status': serializers.CharField()}
        )}, tags=tag
    )
    def post(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role == 'moderator'):
            return Response({"detail": "Доступ запрещён"}, status=status.HTTP_403_FORBIDDEN)
        desired_status = request.data.get('status')
        if desired_status not in dict(Surveys.STATUS_CHOICES):
            return Response({"detail": "Недопустимый статус"}, status=status.HTTP_400_BAD_REQUEST)
        survey.status = desired_status
        survey.save()
        return Response({'survey_id': survey.survey_id, 'status': survey.status},
                        status=status.HTTP_200_OK)


class AvailableSurveysView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(summary="Доступные опросы для респондента",
                   responses={200: SurveyDetailSerializer(many=True)}, tags=tag)
    def get(self, request):
        now = timezone.now()
        qs = Surveys.objects.filter(status='active').filter(
            models.Q(date_finished__isnull=True) | models.Q(date_finished__gt=now)
        )
        available = []
        for s in qs:
            if s.max_residents:
                respondents_count = RespondentAnswers.objects.filter(
                    survey_question__survey=s
                ).values('respondent').distinct().count()
                if respondents_count >= s.max_residents:
                    continue
            available.append(s)
        return Response(SurveyDetailSerializer(available, many=True).data,
                        status=status.HTTP_200_OK)


# ------------------- ИМПОРТ / ЭКСПОРТ -------------------

class ExportSurveyQuestionsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Экспорт вопросов в CSV или XLSX",
        responses={200: None},
        tags=tag
    )
    def get(self, request, survey_id: int, format_type: str):
        survey = Surveys.objects.get(pk=survey_id)
        questions = Questions.objects.filter(survey_questions__survey=survey)

        if format_type == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = f'attachment; filename="survey_{survey_id}_questions.csv"'

            writer = csv.writer(response)
            writer.writerow(["question_id", "text_question", "type_question", "extra_data"])
            for q in questions:
                writer.writerow([q.pk, q.text_question, q.type_question, q.extra_data])

            return response

        elif format_type == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Questions"
            ws.append(["question_id", "text_question", "type_question", "extra_data"])
            for q in questions:
                ws.append([q.pk, q.text_question, q.type_question, str(q.extra_data)])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="survey_{survey_id}_questions.xlsx"'
            wb.save(response)
            return response

        return Response({"detail": "Неподдерживаемый формат"}, status=status.HTTP_400_BAD_REQUEST)


class ImportSurveyQuestionsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Импорт вопросов из CSV или XLSX",
        request=None,
        responses={201: None},
        tags=tag
    )
    def post(self, request, survey_id: int, format_type: str):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Нет файла"}, status=status.HTTP_400_BAD_REQUEST)

        survey = Surveys.objects.get(pk=survey_id)

        created = []
        if format_type == "csv":
            decoded_file = file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)
            for row in reader:
                q = Questions.objects.create(
                    text_question=row["text_question"],
                    type_question=row["type_question"],
                    extra_data=row.get("extra_data", "{}"),
                )
                SurveyQuestions.objects.create(survey=survey, question=q)
                created.append(q.pk)

        elif format_type == "xlsx":
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                _, text_question, type_question, extra_data = row
                q = Questions.objects.create(
                    text_question=text_question,
                    type_question=type_question,
                    extra_data=extra_data or "{}",
                )
                SurveyQuestions.objects.create(survey=survey, question=q)
                created.append(q.pk)

        else:
            return Response({"detail": "Неподдерживаемый формат"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"created_questions": created}, status=status.HTTP_201_CREATED)


class MySurveyProgressView(APIView):
    """
    Получение всех опросов, в которых участвовал респондент, со статусами.
    """
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Мои пройденные/активные опросы",
        responses={200: RespondentSurveyStatusSerializer(many=True)},
        tags=['Опросы']
    )
    def get(self, request):
        user = request.user
        statuses = RespondentSurveyStatus.objects.filter(respondent=user).select_related('survey')
        serializer = RespondentSurveyStatusSerializer(statuses, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class SurveyProgressUpdateView(APIView):
    """
    Обновить или создать статус опроса для респондента.
    Если статус изменяется на 'completed', можно указать оценку (`score`).
    """
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Изменить/создать статус опроса (пройден/в процессе)",
        description=(
            "Позволяет пользователю обновить статус прохождения опроса.\n\n"
            "**Пример:**\n"
            "- Если статус `in_progress` — обновляется только состояние.\n"
            "- Если статус `completed` — необходимо (или можно) добавить поле `score` (0.0–1.0)."
        ),
        request=inline_serializer(
            name="SurveyProgressUpdateRequest",
            fields={
                'status': serializers.ChoiceField(
                    choices=['in_progress', 'completed'],
                    help_text="Статус прохождения опроса ('in_progress' или 'completed')"
                ),
                'score': serializers.FloatField(
                    required=False,
                    min_value=0.0,
                    max_value=1.0,
                    help_text="Оценка прохождения (0.0–1.0), требуется при статусе 'completed'"
                )
            }
        ),
        responses={200: RespondentSurveyStatusSerializer},
        tags=['Опросы']
    )
    def post(self, request, survey_id: int):
        user = request.user
        survey = get_object_or_404(Surveys, pk=survey_id)

        status_value = request.data.get('status')
        score_value = request.data.get('score')

        # Проверка допустимости статуса
        if status_value not in dict(RespondentSurveyStatus.STATUS_CHOICES):
            return Response({"detail": "Недопустимый статус"}, status=status.HTTP_400_BAD_REQUEST)

        # Если статус "completed", проверяем наличие оценки
        if status_value == 'completed':
            if score_value is None:
                return Response(
                    {"detail": "Поле 'score' обязательно при статусе 'completed'"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            try:
                score_value = float(score_value)
                if not (0.0 <= score_value <= 1.0):
                    raise ValueError
            except ValueError:
                return Response(
                    {"detail": "Оценка должна быть числом от 0.0 до 1.0"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            # Для статуса "in_progress" оценка сбрасывается
            score_value = None

        # Обновление или создание записи
        record, created = RespondentSurveyStatus.objects.update_or_create(
            respondent=user,
            survey=survey,
            defaults={'status': status_value, 'score': score_value}
        )

        action = "Создан" if created else "Обновлён"
        print(f"🔄 [{user}] {action} статус: {survey.name} → {status_value} (оценка: {score_value})")

        serializer = RespondentSurveyStatusSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)


class RespondentSurveyAnswersView(APIView):
    permission_classes = [IsSurveyParticipantOrAdmin]

    @extend_schema(
        summary="Получить все ответы респондента по конкретному опросу",
        description="Респондент получает только свои ответы, админ и модератор — все.",
        responses={200: inline_serializer(
            name='ОтветыРеспондентаПоОпросу',
            fields={
                'survey_id': serializers.IntegerField(),
                'respondent_id': serializers.IntegerField(),
                'answers': serializers.ListField(child=serializers.DictField())
            }
        )},
        tags=['Опросы']
    )
    def get(self, request, survey_id: int):
        user = request.user
        print(f"[DEBUG] 🔍 Запрос GET /api/surveys/{survey_id}/answers/ от {user} (роль={getattr(user, 'role', None)})")

        survey = get_object_or_404(Surveys, pk=survey_id)

        # Проверяем разрешения на уровне объекта
        self.check_object_permissions(request, survey)

        role = getattr(user, 'role', None)
        if role in ["moderator", "customer"] and survey.creator == user:
            answers = RespondentAnswers.objects.filter(
                survey_question__survey=survey
            ).select_related("respondent", "survey_question__question")
            print(f"[DEBUG] ✅ Администратор/модератор видит {answers.count()} ответов.")
        else:
            answers = RespondentAnswers.objects.filter(
                respondent=user,
                survey_question__survey=survey
            ).select_related("survey_question__question")
            print(f"[DEBUG] ✅ Респондент видит {answers.count()} своих ответов.")

        answers_list = [
            {
                "answer_id": ans.answer_id,
                "question_id": ans.survey_question.question.question_id,
                "question_text": ans.survey_question.question.text_question,
                "type_question": ans.survey_question.question.type_question,
                "text_answer": ans.text_answer,
                "created_at": ans.created_at,
            }
            for ans in answers
        ]

        print(f"[DEBUG] 🧩 Отправляется {len(answers_list)} ответов клиенту.")
        return Response({
            "survey_id": survey.survey_id,
            "respondent_id": user.id,
            "answers": answers_list
        }, status=status.HTTP_200_OK)

class RespondentAnswerDetailView(APIView):
    """
    Получение ответа текущего респондента на конкретный вопрос по его ID.
    """
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Получить ответ респондента на конкретный вопрос",
        description="Доступно только пользователю с ролью respondent, возвращает один ответ (если есть).",
        responses={200: RespondentAnswerDetailSerializer},
        tags=['Ответы респондентов']
    )
    def get(self, request, question_id: int):
        user = request.user
        role = getattr(user, 'role', None)

        print(f"[DEBUG] 🔍 GET /api/questions/{question_id}/my-answer/ от {user} (роль={role})")

        # Проверка авторизации
        if not user.is_authenticated:
            print("[ERROR] ❌ Пользователь не аутентифицирован.")
            return Response({"detail": "Требуется авторизация."}, status=status.HTTP_401_UNAUTHORIZED)

        # Только респонденты имеют доступ
        if role != "respondent":
            print(f"[DENY] 🚫 Доступ запрещён — роль {role} не является респондентом.")
            return Response({"detail": "Только респонденты могут просматривать свои ответы."},
                            status=status.HTTP_403_FORBIDDEN)

        # Проверяем существование вопроса
        question = get_object_or_404(SurveyQuestions, pk=question_id)
        print(f"[DEBUG] ✅ Найден вопрос ID={question_id}: {question.text_question}")

        # Ищем ответ респондента
        answer = RespondentAnswers.objects.filter(
            respondent=user,
            survey_question__question=question
        ).select_related("survey_question__question").first()

        if not answer:
            print(f"[WARN] ⚠️ У пользователя {user} нет ответа на вопрос ID={question_id}")
            return Response({"detail": "Вы не отвечали на этот вопрос."},
                            status=status.HTTP_404_NOT_FOUND)

        print(f"[DEBUG] ✅ Найден ответ ID={answer.answer_id}, текст={answer.text_answer!r}")

        serializer = RespondentAnswerDetailSerializer(answer)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SurveyAddCharacteristicView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Добавить характеристику к опросу",
        description="Создаёт одну связь между опросом и характеристикой с возможным описанием требований.",
        request=inline_serializer(
            name="SurveyAddCharacteristicRequest",
            fields={
                "characteristic_id": serializers.IntegerField(),
                "requirements": serializers.CharField(allow_blank=True, required=False)
            }
        ),
        responses={201: SurveyRequiredCharacteristicSerializer},
        tags=tag
    )
    def post(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        if not (survey.creator == request.user or request.user.role in ["moderator"]):
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)

        char_id = request.data.get("characteristic_id")
        requirements = request.data.get("requirements", "")

        if not isinstance(char_id, int):
            return Response({"detail": "characteristic_id должен быть числом."}, status=status.HTTP_400_BAD_REQUEST)

        obj, created = SurveyRequiredCharacteristics.objects.get_or_create(
            survey=survey,
            characteristic_id=char_id,
            defaults={"requirements": requirements}
        )
        if not created:
            return Response({"detail": "Эта характеристика уже добавлена."}, status=status.HTTP_400_BAD_REQUEST)

        return Response(SurveyRequiredCharacteristicSerializer(obj).data, status=status.HTTP_201_CREATED)


class SurveyCharacteristicsListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Получить все характеристики опроса",
        description="Возвращает список всех характеристик, связанных с данным опросом.",
        responses={200: SurveyRequiredCharacteristicSerializer(many=True)},
        tags=tag
    )
    def get(self, request, survey_id: int):
        survey = get_object_or_404(Surveys, pk=survey_id)
        links = SurveyRequiredCharacteristics.objects.filter(survey=survey)
        serializer = SurveyRequiredCharacteristicSerializer(links, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class SurveyEditCharacteristicView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Редактировать требования характеристики в опросе",
        request=inline_serializer(
            name="SurveyEditCharacteristicRequest",
            fields={"requirements": serializers.CharField(allow_blank=True)}
        ),
        responses={200: SurveyRequiredCharacteristicSerializer},
        tags=tag
    )
    def put(self, request, survey_id: int, link_id: int):
        link = get_object_or_404(SurveyRequiredCharacteristics, pk=link_id, survey_id=survey_id)
        if not (link.survey.creator == request.user or request.user.role in ["moderator"]):
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)

        link.requirements = request.data.get("requirements", "")
        link.save()
        return Response(SurveyRequiredCharacteristicSerializer(link).data, status=status.HTTP_200_OK)


class SurveyDeleteCharacteristicView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        summary="Удалить характеристику из опроса",
        responses={204: None},
        tags=tag
    )
    def delete(self, request, survey_id: int, link_id: int):
        link = get_object_or_404(SurveyRequiredCharacteristics, pk=link_id, survey_id=survey_id)
        if not (link.survey.creator == request.user or request.user.role in ["moderator"]):
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)

        link.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
